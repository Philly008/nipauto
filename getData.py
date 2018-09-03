# coding = utf-8
from openpyxl import load_workbook

wb = load_workbook(r'D:\SVN\07 性能测试脚本\NIPauto\com\预生产环境\康源LIS V3.5.15.0测试计划.xlsx')
normal = wb.get_sheet_by_name('病理')
normal2 = wb.get_sheet_by_name('常规')
#center_items_sheet = wb.get_sheet_by_name('中心检验项目')

def get_pathology_types():
    pathology_types = []
    for r in range(2, normal.max_row + 1):
        pathology_type = normal.cell(row=r, column=2).value
        pathology_types.append(pathology_type)
    return pathology_types

def get_customers():
    customers = []
    for r in range(2, normal.max_row + 1):
        customer = normal.cell(row=r, column=5).value
        customers.append(customer)
    return customers

def get_barcodes():
    #print('最大行数：' + str(normal.max_row))
    #print('最大列数：' + str(normal.max_column))

    barcodes = []
    for r in range(2, normal.max_row + 1):
        barcode = normal.cell(row=r, column=7).value
        barcodes.append(barcode)
    return barcodes

def get_normal_barcodes():
    #print('最大行数：' + str(normal.max_row))
    #print('最大列数：' + str(normal.max_column))

    barcodes = []
    for r in range(2, normal2.max_row + 1):
        barcode = normal2.cell(row=r, column=7).value
        barcodes.append(barcode)
    return barcodes

def get_normal_customers():
    customers = []
    for r in range(2, normal2.max_row + 1):
        customer = normal2.cell(row=r, column=5).value
        customers.append(customer)
    return customers

def get_normal_instruments():
    instruments = []
    for r in range(2, normal2.max_row + 1):
        instrument = normal2.cell(row=r, column=2).value
        instruments.append(instrument)
    return instruments

def get_normal_items():
    items = []
    for r in range(2, normal2.max_row + 1):
        item = normal2.cell(row=r, column=3).value
        items.append(item)
    return items

def get_report_templates():
    templates = []
    for t in range(2, normal2.max_row + 1):
        t = normal2.cell(row=t, column=4).value
        templates.append(t)
    return templates

def get_report_heads():
    heads = []
    for h in range(2, normal2.max_row + 1):
        h = normal2.cell(row=h, column=6).value
        heads.append(h)
    return heads

def get_pathology_templates():
    tps = []
    for tp in range(2, normal.max_row + 1):
        tp = normal.cell(row=tp, column=4).value
        tps.append(tp)
    return tps

def get_pathology_heads():
    pheads = []
    for ph in range(2, normal.max_row + 1):
        ph = normal.cell(row=ph, column=6).value
        pheads.append(ph)
    return pheads

def get_pathology_bldl():
    bldls = []
    for bldl in range(2, normal.max_row + 1):
        bldl = normal.cell(row=bldl, column=8).value
        bldls.append(bldl)
    return bldls



if __name__ == '__main__':
    print('出现异常的条码号为：' + str(222) + '\n' + str(33))

    for i in range(0, 10):
        #print(get_center_items()[i].value)

        #print(get_pathology_types()[i])
        #print(get_barcodes()[i])
        # print(get_instuments()[i])
        # print(get_items()[i])
        # print(normal.max_row)
        # print(get_normal_barcodes()[i])
        # print(get_normal_customers()[i])
        # print(get_normal_instruments()[i])
        # print(get_normal_items()[i])

        print(get_pathology_bldl()[i])


