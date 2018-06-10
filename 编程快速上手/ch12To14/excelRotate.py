# -*- coding:utf-8 -*-
import openpyxl
wb = openpyxl.load_workbook('multiplicationTable.xlsx')
sheet1 = wb.active
nb = openpyxl.Workbook()
sheet2 = nb.active

rowmax = sheet1.max_row
clmmax = sheet1.max_column
# 二维列表创建方法
sheetdata = [[0 for col in range(clmmax)] for row in range(rowmax)]
for i in range(1, rowmax+1):
    for j in range(1, clmmax+1):
        sheetdata[i-1][j-1] = sheet1.cell(row = i, column = j).value
        sheet2.cell(row=j, column=i).value = sheetdata[i-1][j-1]

nb.save('单元格翻转.xlsx')