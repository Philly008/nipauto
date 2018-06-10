import openpyxl
wb = openpyxl.load_workbook('文本文件到表格.xlsx')
sheet = wb.active
rowmax = sheet.max_row
clnmax = sheet.max_column

for i in range(1, clnmax + 1):
    f = open('file' + str(i) + '.txt', 'w')     # 创建新的 txt 文件
    for j in range(1, rowmax + 1):
        content = sheet.cell(row=j, column=i).value
        if content != None:
            f.write(content)
    f.close()