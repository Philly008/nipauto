# -*- coding:utf-8 -*-
from trace import Trace
import openpyxl
import sys
from openpyxl.styles import Font

# script first argument
max_num = int(sys.argv[1])

# init excel info
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'multiplicationTable'
font = Font(bold=True)

# init first row ... 1 2 3 4 ...
for column in range(1, max_num + 1):
    multiplier_cell = sheet.cell(row = 1, column = column + 1)
    multiplier_cell.font = font
    multiplier_cell.value = column

for row in range(1, max_num + 1):
    multiplican_cell = sheet.cell(row = row + 1, column =1)
    multiplican_cell.font = font
    multiplican_cell.value = row

# solve by dynamic programming
for row in range(1, max_num + 1):
    for column in range(1, max_num + 1):
        # each column, first row
        multiplier_cell = sheet.cell(row=1, column=column + 1)
        multiplican_cell = sheet.cell(row=row+1, column=1)
        # result[i][j] = result[1][i] * result[i][1]
        result_cell = sheet.cell(row=row + 1, column = column + 1)
        multiplier_value = multiplier_cell.value
        multiplicand_value = multiplican_cell.value
        result_cell.value = multiplier_value * multiplicand_value

wb.save('multiplicationTable.xlsx')